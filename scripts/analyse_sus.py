#!/usr/bin/env python3
"""
System Usability Scale analysis - chapter 3.3 / 5.6

Scores the ten SUS items from the questionnaire export exactly as specified by
Brooke (38): for items 1, 3, 5, 7 and 9 the contribution is the scale position
minus one; for items 2, 4, 6, 8 and 10 it is five minus the scale position; the
sum is multiplied by 2.5 to give a score from 0 to 100. Individual item scores
are not interpreted on their own, as Brooke notes they are not meaningful in
isolation.

One item was left blank by one respondent. Brooke's instruction for that case
is that the respondent "should mark the centre point of the scale", so the
missing value is imputed as Neutral (3) rather than dropping the respondent.
The effect either way is reported so the choice is transparent.

Comparison figures come from Vlachogianni and Tselios (39), a systematic review
of 104 studies (N = 170 surveys).

usage:
    python scripts/analyse_sus.py
"""
from __future__ import annotations

import csv
import json
import statistics
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required: pip install openpyxl")
    raise SystemExit(1)

XLSX = Path(r"C:\Users\Ataur Rahman\Downloads"
            r"\U-Pal_ A Bilingual Welsh_English Chatbot for Student Support Services (Student 18-25)(1-50).xlsx")
OUT_DIR = (Path(__file__).resolve().parent.parent
           / "benchmark_results" / "5.6_usability")
OUT_DIR.mkdir(parents=True, exist_ok=True)

LIKERT = {"strongly disagree": 1, "disagree": 2, "neutral": 3,
          "agree": 4, "strongly agree": 5}

SUS_COLS = list(range(16, 26))          # ten SUS items, in order
POSITIVE = {0, 2, 4, 6, 8}              # zero-based indices of items 1,3,5,7,9

COL_YEAR, COL_WELSH, COL_FREQ, COL_PRIOR = 8, 9, 10, 11
TASK_COLS = [12, 13, 14, 15]
BILINGUAL_COLS = list(range(26, 32))
OPEN_COLS = [32, 33, 34]

# Vlachogianni & Tselios (39)
BENCHMARKS = [
    ("All educational technology (N=170)", 70.09, 12.98),
    ("Multimedia", 76.43, 9.45),
    ("Mobile applications", 73.62, 13.49),
    ("Affective tutoring systems", 68.87, 7.30),
    ("Internet platforms", 66.25, 12.42),
    ("University websites", 63.82, 16.52),
]


def main() -> int:
    ws = load_workbook(XLSX, data_only=True)["Sheet1"]
    head = {c: str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)}
    rows = []
    imputed = 0

    for r in range(2, ws.max_row + 1):
        def cell(c):
            v = ws.cell(row=r, column=c).value
            return str(v).strip() if v is not None else ""

        raw = [cell(c) for c in SUS_COLS]
        positions, miss = [], 0
        for v in raw:
            key = v.lower()
            if key in LIKERT:
                positions.append(LIKERT[key])
            else:
                positions.append(3)     # Brooke: centre point when unanswered
                miss += 1
        imputed += miss

        contrib = [(p - 1) if i in POSITIVE else (5 - p) for i, p in enumerate(positions)]
        sus = sum(contrib) * 2.5
        rows.append({
            "id": cell(1), "sus": sus, "imputed": miss,
            "year": cell(COL_YEAR), "welsh": cell(COL_WELSH),
            "freq": cell(COL_FREQ), "prior": cell(COL_PRIOR),
            "tasks": [cell(c) for c in TASK_COLS],
            "bilingual": [cell(c) for c in BILINGUAL_COLS],
            "open": [cell(c) for c in OPEN_COLS],
            "positions": positions,
        })

    scores = [r["sus"] for r in rows]
    n = len(scores)
    mean = statistics.mean(scores)
    sd = statistics.stdev(scores)
    se = sd / (n ** 0.5)
    ci = (mean - 1.96 * se, mean + 1.96 * se)

    print("=" * 74)
    print("SYSTEM USABILITY SCALE")
    print("=" * 74)
    print(f"respondents            : {n}")
    print(f"imputed item responses : {imputed} (centre point, per Brooke)")
    print(f"mean SUS               : {mean:.2f}")
    print(f"standard deviation     : {sd:.2f}")
    print(f"median                 : {statistics.median(scores):.2f}")
    print(f"range                  : {min(scores):.1f} - {max(scores):.1f}")
    print(f"95% CI of the mean     : {ci[0]:.2f} to {ci[1]:.2f}")

    # sensitivity: drop the respondent with the imputed item
    kept = [r["sus"] for r in rows if r["imputed"] == 0]
    print(f"excluding the imputed response: n={len(kept)}, "
          f"mean {statistics.mean(kept):.2f}, sd {statistics.stdev(kept):.2f}")

    print("\nComparison with Vlachogianni & Tselios (39)")
    print(f"{'Category':<38}{'Mean':>8}{'SD':>8}{'U-Pal diff':>12}")
    for label, m, s in BENCHMARKS:
        print(f"{label:<38}{m:>8.2f}{s:>8.2f}{mean - m:>+12.2f}")

    def group(key, label):
        print(f"\nBy {label}")
        buckets: dict[str, list[float]] = {}
        for r in rows:
            buckets.setdefault(r[key] or "(blank)", []).append(r["sus"])
        for k in sorted(buckets, key=lambda x: -len(buckets[x])):
            v = buckets[k]
            sd_v = f"{statistics.stdev(v):.2f}" if len(v) > 1 else "-"
            print(f"  {k:<28}n={len(v):<4}mean {statistics.mean(v):>6.2f}  sd {sd_v}")

    group("welsh", "Welsh language ability")
    group("year", "year of study")
    group("prior", "prior use of a university support chatbot")
    group("freq", "general chatbot use")

    print("\nTask completion (questionnaire items 6-9)")
    for i, c in enumerate(TASK_COLS):
        counts = Counter(r["tasks"][i] for r in rows)
        print(f"  {head[c][:58]}")
        for k, v in counts.most_common():
            print(f"      {v:>3}  {k[:60]}")

    print("\nBilingual and trust items (mean of 1-5 Likert)")
    for i, c in enumerate(BILINGUAL_COLS):
        vals = [LIKERT.get(r["bilingual"][i].lower()) for r in rows]
        vals = [v for v in vals if v]
        agree = sum(1 for v in vals if v >= 4)
        print(f"  {head[c][:62]:<64} mean {statistics.mean(vals):.2f}  "
              f"agree+ {agree}/{len(vals)} ({agree/len(vals):.0%})")

    # exports
    with open(OUT_DIR / "sus_per_respondent.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Respondent", "SUS score", "Imputed items", "Year",
                    "Welsh ability", "Prior chatbot use"]
                   + [f"Item {i+1}" for i in range(10)])
        for r in rows:
            w.writerow([r["id"], r["sus"], r["imputed"], r["year"], r["welsh"],
                        r["prior"]] + r["positions"])

    with open(OUT_DIR / "sus_summary.json", "w", encoding="utf-8") as f:
        json.dump({"n": n, "mean": mean, "sd": sd,
                   "median": statistics.median(scores),
                   "min": min(scores), "max": max(scores), "ci95": ci,
                   "imputed_items": imputed,
                   "scores": scores}, f, indent=2)

    with open(OUT_DIR / "open_ended_responses.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Respondent", "SUS", "Welsh ability"] + [head[c][:60] for c in OPEN_COLS])
        for r in rows:
            if any(x and x.lower() != "none" for x in r["open"]):
                w.writerow([r["id"], r["sus"], r["welsh"]] + r["open"])

    print(f"\nwrote {(OUT_DIR / 'sus_per_respondent.csv').name}")
    print(f"wrote {(OUT_DIR / 'sus_summary.json').name}")
    print(f"wrote {(OUT_DIR / 'open_ended_responses.csv').name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
